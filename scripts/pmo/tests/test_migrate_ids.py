"""Tests for migrate-ids subcommand and lib/migrate.py."""
from __future__ import annotations

import shutil
from pathlib import Path
from unittest.mock import patch

import openpyxl
import pytest

from lib.migrate import compute_id_assignments, scan_rows, write_id_cells


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------


def _make_workbook(tmp_path: Path, rows: list[list]) -> Path:
    """Create a minimal .xlsx with given rows starting at row 1, column A onward."""
    p = tmp_path / "test.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    for r_idx, row in enumerate(rows, start=1):
        for c_idx, val in enumerate(row, start=1):
            ws.cell(row=r_idx, column=c_idx, value=val)
    wb.save(p)
    return p


# ---------------------------------------------------------------------------
# Unit: compute_id_assignments
# ---------------------------------------------------------------------------


def test_all_empty_ids_starts_at_001() -> None:
    """All-None id list → T-001, T-002, ..."""
    ids: list[str | None] = [None, None, None]
    result = compute_id_assignments(ids)
    assert result == [(0, "T-001"), (1, "T-002"), (2, "T-003")]


def test_mixed_existing_and_empty_ids() -> None:
    """Existing T-001 present → empty slots get T-002, T-003."""
    ids: list[str | None] = ["T-001", None, None]
    result = compute_id_assignments(ids)
    assert result == [(1, "T-002"), (2, "T-003")]


def test_no_empty_ids_returns_empty_list() -> None:
    """No empty entries → empty assignment list."""
    ids: list[str | None] = ["T-001", "T-002", "T-003"]
    result = compute_id_assignments(ids)
    assert result == []


def test_garbage_ids_are_skipped_in_max_detection() -> None:
    """Non-T-NNN ids are ignored when computing max; empty slots start at T-001."""
    ids: list[str | None] = ["GARBAGE", None]
    result = compute_id_assignments(ids)
    assert result == [(1, "T-001")]


def test_whitespace_only_id_treated_as_empty() -> None:
    """Whitespace-only id string counts as empty."""
    ids: list[str | None] = ["T-001", "   ", None]
    result = compute_id_assignments(ids)
    assert result == [(1, "T-002"), (2, "T-003")]


# ---------------------------------------------------------------------------
# Integration: scan_rows + write_id_cells (case 1: all empty)
# ---------------------------------------------------------------------------


def test_scan_and_assign_all_empty(tmp_path: Path) -> None:
    """Workbook with empty id column → assignments cover every data row."""
    # col A = id (empty), col B = name
    p = _make_workbook(tmp_path, [["", "Task A"], ["", "Task B"], ["", "Task C"]])
    rows = scan_rows(
        p,
        sheet="Sheet1",
        data_start_row=1,
        id_column="A",
        all_columns=["A", "B"],
    )
    ids = [r["A"] for r in rows]
    assignments = compute_id_assignments(ids)
    assert len(assignments) == 3
    assert assignments[0] == (0, "T-001")
    assert assignments[1] == (1, "T-002")
    assert assignments[2] == (2, "T-003")


# ---------------------------------------------------------------------------
# Integration: scan_rows + write_id_cells (case 2: mixed existing + empty)
# ---------------------------------------------------------------------------


def test_scan_and_assign_mixed(tmp_path: Path) -> None:
    """T-001 already set; only empty rows get new ids."""
    p = _make_workbook(
        tmp_path, [["T-001", "Task A"], ["", "Task B"], ["", "Task C"]]
    )
    rows = scan_rows(
        p,
        sheet="Sheet1",
        data_start_row=1,
        id_column="A",
        all_columns=["A", "B"],
    )
    ids = [r["A"] for r in rows]
    assignments = compute_id_assignments(ids)
    assert assignments == [(1, "T-002"), (2, "T-003")]


def test_write_id_cells_actually_writes(tmp_path: Path) -> None:
    """write_id_cells persists the assigned ids into the workbook."""
    p = _make_workbook(
        tmp_path, [["T-001", "Task A"], ["", "Task B"]]
    )
    rows = scan_rows(
        p,
        sheet="Sheet1",
        data_start_row=1,
        id_column="A",
        all_columns=["A", "B"],
    )
    ids = [r["A"] for r in rows]
    assignments = compute_id_assignments(ids)
    row_map = [r["row"] for r in rows]
    write_id_cells(p, sheet="Sheet1", id_column="A", assignments=assignments, row_map=row_map)

    # verify by re-reading
    wb = openpyxl.load_workbook(p, data_only=True)
    ws = wb["Sheet1"]
    assert ws.cell(row=1, column=1).value == "T-001"
    assert ws.cell(row=2, column=1).value == "T-002"


# ---------------------------------------------------------------------------
# CLI: --dry-run does not write back (case 3)
# ---------------------------------------------------------------------------


def test_dry_run_does_not_write(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """--dry-run flag: id assignments are printed but not written to the workbook."""
    import sys
    from sync import main

    p = _make_workbook(tmp_path, [["", "Task A"], ["", "Task B"]])
    # Build a minimal pmo.yaml for the project
    pdir = tmp_path / "proj"
    pdir.mkdir()
    (pdir / ".pmo").mkdir()
    (pdir / "pmo.yaml").write_text(
        f"""project:
  name: test
  slug: test
  start: "2026-01-01"
  end: "2026-12-31"
excel:
  file: test.xlsx
  sheet: Sheet1
  header_row: 0
  data_start_row: 1
  id_column: A
  columns:
    - {{col: A, field: id, owner: yaml}}
    - {{col: B, field: name, owner: yaml}}
tasks: []
""",
        encoding="utf-8",
    )
    shutil.copy(p, pdir / "test.xlsx")

    # Record mtime before running
    target = pdir / "test.xlsx"
    mtime_before = target.stat().st_mtime

    # patch project_dir so it resolves to our tmp_path project
    monkeypatch.chdir(tmp_path)

    import lib.migrate as migrate_mod

    original_write = migrate_mod.write_id_cells
    write_called = []

    def _capture(*args, **kwargs):
        write_called.append(True)
        return original_write(*args, **kwargs)

    monkeypatch.setattr(migrate_mod, "write_id_cells", _capture)

    # Patch project_dir to return pdir
    import sync as sync_mod

    monkeypatch.setattr(sync_mod, "project_dir", lambda slug: pdir)

    captured_output: list[str] = []
    monkeypatch.setattr(
        "sys.stdout",
        type("FakeOut", (), {"write": lambda self, s: captured_output.append(s), "flush": lambda self: None})(),
    )

    ret = main(["migrate-ids", "wbs", "--project", "test", "--dry-run"])

    assert ret == 0
    # write_id_cells must NOT be called in dry-run
    assert not write_called
    output = "".join(captured_output)
    # Should mention the assignments
    assert "T-001" in output
    assert "T-002" in output
