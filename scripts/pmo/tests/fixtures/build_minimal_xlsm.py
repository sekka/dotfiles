from __future__ import annotations

import io
import zipfile
from pathlib import Path

import openpyxl

FIXTURE_PATH = Path(__file__).parent / "minimal.xlsm"
DUMMY_VBA = b"\x01CONST_TEST_VBA_"

_VBA_CT_ENTRY = (
    '<Override PartName="/xl/vbaProject.bin"'
    ' ContentType="application/vnd.ms-office.vbaProject"/>'
)

_VBA_REL_ENTRY = (
    '<Relationship Id="rId99" Type="http://schemas.microsoft.com/office/2006/'
    'relationships/vbaProject" Target="vbaProject.bin"/>'
)


def _patch_content_types(xml_bytes: bytes) -> bytes:
    text = xml_bytes.decode()
    if "vbaProject" not in text:
        text = text.replace("</Types>", f"{_VBA_CT_ENTRY}</Types>")
    return text.encode()


def _patch_workbook_rels(xml_bytes: bytes) -> bytes:
    text = xml_bytes.decode()
    if "vbaProject" not in text:
        text = text.replace("</Relationships>", f"{_VBA_REL_ENTRY}</Relationships>")
    return text.encode()


def build() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="T-001")
    ws.cell(row=1, column=2, value="alpha")
    ws.cell(row=2, column=1, value="T-002")
    ws.cell(row=2, column=2, value="beta")

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    src_bytes = buf.read()

    with zipfile.ZipFile(FIXTURE_PATH, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        with zipfile.ZipFile(io.BytesIO(src_bytes)) as src:
            for info in src.infolist():
                data = src.read(info.filename)
                if info.filename == "[Content_Types].xml":
                    data = _patch_content_types(data)
                elif info.filename == "xl/_rels/workbook.xml.rels":
                    data = _patch_workbook_rels(data)
                dst.writestr(info, data)
        dst.writestr("xl/vbaProject.bin", DUMMY_VBA)


if __name__ == "__main__":
    build()
    print(f"Generated: {FIXTURE_PATH} ({FIXTURE_PATH.stat().st_size} bytes)")
