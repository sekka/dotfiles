"""End-to-end tests for the pull/init/doctor CLI."""
from __future__ import annotations

import sys
from pathlib import Path
from types import SimpleNamespace

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).parent.parent))

import sync as sync_mod
from lib.yaml_io import load_yaml


def _make_wbs_xlsx(path: Path) -> None:
    """Create a minimal WBS workbook matching WBS_SCHEMA (header row 4, data start row 6)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WBS"
    headers = [
        "id", "phase_l1", "phase_l2", "name", "assignee",
        "est_days", "start_date", "end_date", "notes", "status",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=4, column=col_idx, value=h)
    ws.cell(row=6, column=1, value="T-001")
    ws.cell(row=6, column=2, value="現状把握")
    ws.cell(row=6, column=4, value="ヒアリング")
    ws.cell(row=6, column=10, value="進行中")
    ws.cell(row=7, column=1, value="T-002")
    ws.cell(row=7, column=4, value="次タスク")
    wb.save(path)


def _setup_project(pdir: Path, *, with_excel: bool = True) -> Path:
    pdir.mkdir(parents=True, exist_ok=True)
    (pdir / "WBS.yaml").write_text(
        'project:\n  name: "T"\n  slug: "t"\n\nexcel:\n  file: "WBS.xlsx"\n\ntasks: []\n',
        encoding="utf-8",
    )
    if with_excel:
        _make_wbs_xlsx(pdir / "WBS.xlsx")
    return pdir


def _args(**overrides) -> SimpleNamespace:
    base = {"project": None, "pdir": None, "file": None, "force": False}
    base.update(overrides)
    return SimpleNamespace(**base)


def test_pull_regenerates_tasks_from_excel(tmp_path):
    pdir = _setup_project(tmp_path / "proj")
    rc = sync_mod.cmd_pull(_args(pdir=pdir))
    assert rc == 0
    doc = load_yaml(pdir / "WBS.yaml")
    ids = [t["id"] for t in doc["tasks"]]
    assert ids == ["T-001", "T-002"]
    assert doc["tasks"][0]["phase_l1"] == "現状把握"
    assert doc["tasks"][0]["status"] == "進行中"


def test_pull_overwrites_previous_tasks(tmp_path):
    """If YAML had stale tasks not in Excel, pull must drop them."""
    pdir = _setup_project(tmp_path / "proj")
    (pdir / "WBS.yaml").write_text(
        'project:\n  name: "T"\n  slug: "t"\n\nexcel:\n  file: "WBS.xlsx"\n\n'
        'tasks:\n  - id: T-999\n    name: ghost\n',
        encoding="utf-8",
    )
    rc = sync_mod.cmd_pull(_args(pdir=pdir))
    assert rc == 0
    doc = load_yaml(pdir / "WBS.yaml")
    ids = [t["id"] for t in doc["tasks"]]
    assert "T-999" not in ids
    assert ids == ["T-001", "T-002"]


def test_pull_preserves_project_header(tmp_path):
    pdir = _setup_project(tmp_path / "proj")
    (pdir / "WBS.yaml").write_text(
        'project:\n  name: "Custom"\n  slug: "custom"\n  owner: "alice"\n\n'
        'excel:\n  file: "WBS.xlsx"\n\ntasks: []\n',
        encoding="utf-8",
    )
    rc = sync_mod.cmd_pull(_args(pdir=pdir))
    assert rc == 0
    doc = load_yaml(pdir / "WBS.yaml")
    assert doc["project"]["name"] == "Custom"
    assert doc["project"]["owner"] == "alice"
    assert doc["excel"]["file"] == "WBS.xlsx"


def test_pull_aborts_when_yaml_missing(tmp_path, capsys):
    pdir = tmp_path / "proj"
    pdir.mkdir()
    rc = sync_mod.cmd_pull(_args(pdir=pdir))
    assert rc == 2
    err = capsys.readouterr().err
    assert "WBS.yaml not found" in err


def test_pull_aborts_when_excel_missing(tmp_path, capsys):
    pdir = _setup_project(tmp_path / "proj", with_excel=False)
    rc = sync_mod.cmd_pull(_args(pdir=pdir))
    assert rc == 2
    err = capsys.readouterr().err
    assert "excel file not found" in err


def test_init_creates_skeleton(tmp_path, monkeypatch):
    monkeypatch.setattr(Path, "home", classmethod(lambda cls: tmp_path))
    rc = sync_mod.cmd_init(_args(project="new", file=None, force=False))
    assert rc == 0
    pmo = tmp_path / "prj" / "new" / "WBS.yaml"
    assert pmo.exists()
    doc = load_yaml(pmo)
    assert doc["project"]["slug"] == "new"
    assert doc["excel"]["file"] == "WBS.xlsx"
    assert doc["tasks"] == []


def test_init_refuses_existing_without_force(tmp_path, monkeypatch, capsys):
    monkeypatch.setattr(Path, "home", classmethod(lambda cls: tmp_path))
    sync_mod.cmd_init(_args(project="new"))
    rc = sync_mod.cmd_init(_args(project="new"))
    assert rc == 1
    assert "already exists" in capsys.readouterr().err


def test_doctor_passes_on_clean_yaml(tmp_path, capsys):
    pdir = _setup_project(tmp_path / "proj")
    sync_mod.cmd_pull(_args(pdir=pdir))
    rc = sync_mod.cmd_doctor(_args(pdir=pdir))
    assert rc == 0
    assert "no issues" in capsys.readouterr().out


def test_doctor_detects_duplicate_ids(tmp_path, capsys):
    pdir = tmp_path / "proj"
    pdir.mkdir()
    (pdir / "WBS.yaml").write_text(
        'project:\n  name: t\n  slug: t\nexcel:\n  file: WBS.xlsx\n'
        'tasks:\n  - id: T-001\n    name: a\n  - id: T-001\n    name: b\n',
        encoding="utf-8",
    )
    rc = sync_mod.cmd_doctor(_args(pdir=pdir))
    assert rc == 1
    assert "duplicate id" in capsys.readouterr().out
