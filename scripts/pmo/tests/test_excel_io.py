from pathlib import Path

import openpyxl

from lib.excel_io import _resolve_value, read_rows


def make_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "WBS"
    headers = [
        "id", "phase_l1", "phase_l2", "name", "assignee",
        "est_days", "start_date", "end_date", "notes", "status",
    ]
    for col_idx, h in enumerate(headers, start=1):
        ws.cell(row=6, column=col_idx, value=h)
    ws.cell(row=7, column=1, value="T-001")
    ws.cell(row=7, column=2, value="現状把握")
    ws.cell(row=7, column=4, value="ヒアリング")
    ws.cell(row=7, column=10, value="進行中")
    ws.cell(row=8, column=1, value="T-002")
    ws.cell(row=8, column=4, value="次タスク")
    wb.save(path)


def test_resolve_value_formula_with_computed_result():
    assert _resolve_value("=WORKDAY(A1,5)", "2026-05-20") == "2026-05-20"


def test_resolve_value_formula_with_no_computed_result():
    assert _resolve_value("=IF(G6=\"\",WORKDAY(B1,5),G6)", None) is None


def test_resolve_value_plain_value_uses_structural():
    assert _resolve_value("T-001", None) == "T-001"
    assert _resolve_value(None, None) is None
    assert _resolve_value(42, None) == 42


def test_read_rows_returns_dicts_keyed_by_column_letter(tmp_path):
    xlsx = tmp_path / "wbs.xlsx"
    make_workbook(xlsx)
    rows = read_rows(
        xlsx, sheet="WBS", data_start_row=7,
        columns=["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"],
    )
    assert len(rows) == 2
    assert rows[0]["A"] == "T-001"
    assert rows[0]["B"] == "現状把握"
    assert rows[0]["J"] == "進行中"
    assert rows[1]["A"] == "T-002"
    assert rows[1]["D"] == "次タスク"


def test_read_rows_stops_at_first_empty_id(tmp_path):
    xlsx = tmp_path / "wbs.xlsx"
    make_workbook(xlsx)
    wb = openpyxl.load_workbook(xlsx)
    ws = wb["WBS"]
    ws.cell(row=10, column=1, value="T-003")  # row 9 empty, row 10 has data
    wb.save(xlsx)
    rows = read_rows(
        xlsx, sheet="WBS", data_start_row=7,
        columns=["A", "B", "C", "D", "E", "F", "G", "H", "I", "J"],
        id_column="A",
    )
    assert len(rows) == 2
