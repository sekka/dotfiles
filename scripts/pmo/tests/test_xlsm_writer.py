"""TDD tests for lib/xlsm_writer.py

Uses zipfile + lxml approach to write Excel cells without destroying
extLst, customXml, sharedStrings, and other parts that openpyxl corrupts.
"""
from __future__ import annotations

import datetime
import io
import os
import zipfile
from pathlib import Path
from typing import Any

import openpyxl
import pytest
from lxml import etree

# Module under test (will not exist until Step 3)
from lib.xlsm_writer import append_rows, write_cells

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
NS_R = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
NS_X14 = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
NS_MC = "http://schemas.openxmlformats.org/markup-compatibility/2006"


# ---------------------------------------------------------------------------
# Fixture helpers
# ---------------------------------------------------------------------------


def _build_base_xlsx_bytes() -> bytes:
    """Build a minimal xlsx in memory via openpyxl and return raw bytes."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.cell(row=1, column=1, value="T-001")
    ws.cell(row=1, column=2, value="alpha")
    ws.cell(row=2, column=1, value="T-002")
    ws.cell(row=2, column=2, value=42)
    ws.cell(row=3, column=1, value="T-003")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def _inject_extlst_into_sheet_xml(sheet_xml: bytes) -> bytes:
    """Inject an x14:dataValidations extLst block into the sheet XML.

    Adds xmlns:x14 and xmlns:mc on the root element (not just child) so
    lxml retains those namespace declarations after any parse→serialize round-trip.
    lxml only preserves a namespace declaration on the root if it is declared there;
    if it only appears on a child element it may be re-scoped or dropped.
    """
    root = etree.fromstring(sheet_xml)

    # lxml does not allow modifying nsmap in place. Re-create the root element
    # with explicit nsmap so x14 and mc are declared at the document root.
    existing_nsmap = dict(root.nsmap)
    existing_nsmap["x14"] = NS_X14
    existing_nsmap["mc"] = NS_MC

    new_root = etree.Element(root.tag, nsmap=existing_nsmap)
    for k, v in root.attrib.items():
        new_root.set(k, v)
    for child in root:
        new_root.append(child)

    # Add mc:Ignorable on root so MC namespace is actively used
    new_root.set(f"{{{NS_MC}}}Ignorable", "x14")

    # Build extLst with x14 content
    ext_lst = etree.SubElement(new_root, f"{{{NS}}}extLst")
    ext = etree.SubElement(ext_lst, f"{{{NS}}}ext")
    ext.set("uri", "{CCE6A557-97BC-4b89-B988-1F9FDA3B8994}")
    dv_list = etree.SubElement(ext, f"{{{NS_X14}}}dataValidations")
    dv_list.set("count", "1")
    dv = etree.SubElement(dv_list, f"{{{NS_X14}}}dataValidation")
    dv.set("type", "list")
    dv.set("allowBlank", "1")
    dv.set("showDropDown", "1")
    dv.set("sqref", "B1:B10")
    formula = etree.SubElement(dv, f"{{{NS_X14}}}formula1")
    formula.text = '"yes,no"'

    return etree.tostring(new_root, xml_declaration=True, encoding="UTF-8", standalone=True)


@pytest.fixture()
def minimal_xlsm(tmp_path: Path) -> Path:
    """Minimal xlsm with no VBA, suitable for basic cell read/write tests."""
    xlsm = tmp_path / "minimal.xlsm"
    base = _build_base_xlsx_bytes()

    with zipfile.ZipFile(xlsm, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        with zipfile.ZipFile(io.BytesIO(base)) as src:
            for info in src.infolist():
                dst.writestr(info, src.read(info.filename))
    return xlsm


@pytest.fixture()
def xlsm_with_extlst(tmp_path: Path) -> Path:
    """xlsm with extLst in sheet1, x14 namespace, and customXml part."""
    xlsm = tmp_path / "with_extlst.xlsm"
    base = _build_base_xlsx_bytes()

    custom_xml_content = b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n<root><item>sharepoint-metadata</item></root>'

    with zipfile.ZipFile(xlsm, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        with zipfile.ZipFile(io.BytesIO(base)) as src:
            for info in src.infolist():
                data = src.read(info.filename)
                if info.filename.startswith("xl/worksheets/sheet"):
                    data = _inject_extlst_into_sheet_xml(data)
                elif info.filename == "[Content_Types].xml":
                    # Inject customXml Override
                    ct_text = data.decode()
                    override = '<Override PartName="/customXml/item1.xml" ContentType="application/xml"/>'
                    if override not in ct_text:
                        ct_text = ct_text.replace("</Types>", f"{override}</Types>")
                    data = ct_text.encode()
                dst.writestr(info, data)
        # Add customXml
        dst.writestr("customXml/item1.xml", custom_xml_content)

    return xlsm


# ---------------------------------------------------------------------------
# Step 1: write_cells — existing cell updates
# ---------------------------------------------------------------------------


def test_write_cells_updates_existing_string_cell(minimal_xlsm: Path) -> None:
    write_cells(minimal_xlsm, "Sheet1", [(1, "A", "UPDATED")])
    wb = openpyxl.load_workbook(minimal_xlsm, data_only=True)
    assert wb["Sheet1"]["A1"].value == "UPDATED"


def test_write_cells_updates_existing_number_cell(minimal_xlsm: Path) -> None:
    write_cells(minimal_xlsm, "Sheet1", [(2, "B", 99)])
    wb = openpyxl.load_workbook(minimal_xlsm, data_only=True)
    assert wb["Sheet1"]["B2"].value == 99


def test_write_cells_creates_new_cell_in_existing_row(minimal_xlsm: Path) -> None:
    """Row 3 has only A3. Writing C3 should insert a new <c r="C3"> in sorted order."""
    write_cells(minimal_xlsm, "Sheet1", [(3, "C", "NEW")])
    wb = openpyxl.load_workbook(minimal_xlsm, data_only=True)
    assert wb["Sheet1"]["C3"].value == "NEW"
    assert wb["Sheet1"]["A3"].value == "T-003"  # existing cell preserved

    # Verify sorted order in raw XML
    with zipfile.ZipFile(minimal_xlsm) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        xml = zf.read(sheet_names[0])
    root = etree.fromstring(xml)
    rows = root.findall(f".//{{{NS}}}row")
    row3 = next(r for r in rows if r.get("r") == "3")
    cell_refs = [c.get("r") for c in row3.findall(f"{{{NS}}}c")]
    # Should be sorted: A3 before C3
    assert cell_refs == sorted(cell_refs)


def test_write_cells_preserves_existing_cell_style(minimal_xlsm: Path) -> None:
    """The s= (style index) attribute must not be destroyed when updating value.

    Note: t= may legitimately change (e.g., sharedString 's' -> 'inlineStr')
    when the encoding method changes; only s= is guaranteed to survive.
    """
    # First, inject a style attribute onto A1 in the raw XML
    with zipfile.ZipFile(minimal_xlsm) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        sheet_name = sheet_names[0]
        xml = zf.read(sheet_name)

    root = etree.fromstring(xml)
    rows = root.findall(f".//{{{NS}}}row")
    row1 = next(r for r in rows if r.get("r") == "1")
    a1 = next(c for c in row1.findall(f"{{{NS}}}c") if c.get("r") == "A1")
    a1.set("s", "3")  # inject style index
    patched = etree.tostring(root, xml_declaration=True, encoding="UTF-8", standalone=True)

    # Write back the patched XML
    all_entries: list[tuple[str, bytes]] = []
    with zipfile.ZipFile(minimal_xlsm) as zf:
        for info in zf.infolist():
            if info.filename == sheet_name:
                all_entries.append((info.filename, patched))
            else:
                all_entries.append((info.filename, zf.read(info.filename)))
    tmp = minimal_xlsm.with_suffix(".tmp.xlsm")
    with zipfile.ZipFile(tmp, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        for fname, data in all_entries:
            dst.writestr(fname, data)
    os.replace(tmp, minimal_xlsm)

    # Now run writer and check s= is preserved
    write_cells(minimal_xlsm, "Sheet1", [(1, "A", "STYLE_TEST")])

    with zipfile.ZipFile(minimal_xlsm) as zf:
        xml = zf.read(sheet_name)
    root = etree.fromstring(xml)
    rows = root.findall(f".//{{{NS}}}row")
    row1 = next(r for r in rows if r.get("r") == "1")
    a1 = next(c for c in row1.findall(f"{{{NS}}}c") if c.get("r") == "A1")
    assert a1.get("s") == "3", "style index s= must survive a value update"


def test_write_cells_none_deletes_cell(minimal_xlsm: Path) -> None:
    write_cells(minimal_xlsm, "Sheet1", [(1, "A", None)])
    with zipfile.ZipFile(minimal_xlsm) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        xml = zf.read(sheet_names[0])
    root = etree.fromstring(xml)
    rows = root.findall(f".//{{{NS}}}row")
    row1_cells = []
    for r in rows:
        if r.get("r") == "1":
            row1_cells = [c.get("r") for c in r.findall(f"{{{NS}}}c")]
    assert "A1" not in row1_cells, "None value should delete the cell element"


# ---------------------------------------------------------------------------
# Step 2: append_rows
# ---------------------------------------------------------------------------


def test_append_rows_appends_at_end(minimal_xlsm: Path) -> None:
    """New rows are added after the last occupied row."""
    new_rows = [{"A": "T-004", "B": "gamma"}, {"A": "T-005", "B": "delta"}]
    append_rows(minimal_xlsm, "Sheet1", new_rows)
    wb = openpyxl.load_workbook(minimal_xlsm, data_only=True)
    ws = wb["Sheet1"]
    # Existing rows 1-3 remain, new rows at 4-5 (row 3 has only A, so max_row=3)
    assert ws["A4"].value == "T-004"
    assert ws["B4"].value == "gamma"
    assert ws["A5"].value == "T-005"
    assert ws["B5"].value == "delta"
    # Existing preserved
    assert ws["A1"].value == "T-001"


def test_append_rows_returns_first_appended_row(minimal_xlsm: Path) -> None:
    """Return value must be the 1-based row number where insertion started."""
    new_rows = [{"A": "T-NEW"}]
    result = append_rows(minimal_xlsm, "Sheet1", new_rows)
    # minimal has 3 rows (1,2,3), so appended at row 4
    assert result == 4


def test_append_rows_with_start_row(minimal_xlsm: Path) -> None:
    """start_row parameter overrides auto-detect."""
    result = append_rows(minimal_xlsm, "Sheet1", [{"A": "FORCED"}], start_row=10)
    assert result == 10
    wb = openpyxl.load_workbook(minimal_xlsm, data_only=True)
    assert wb["Sheet1"]["A10"].value == "FORCED"


# ---------------------------------------------------------------------------
# Step 3: round-trip preservation
# ---------------------------------------------------------------------------


def test_round_trip_preserves_extlst(xlsm_with_extlst: Path) -> None:
    """After write_cells, the extLst block in sheet XML must still be present."""
    write_cells(xlsm_with_extlst, "Sheet1", [(1, "A", "MODIFIED")])

    with zipfile.ZipFile(xlsm_with_extlst) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        xml = zf.read(sheet_names[0])

    root = etree.fromstring(xml)
    ext_lst = root.find(f"{{{NS}}}extLst")
    assert ext_lst is not None, "extLst must survive write_cells"
    # Check x14:dataValidations child is still present
    dv_list = ext_lst.find(f".//{{{NS_X14}}}dataValidations")
    assert dv_list is not None, "x14:dataValidations inside extLst must survive"


def test_round_trip_preserves_customxml(xlsm_with_extlst: Path) -> None:
    """customXml/item1.xml must be present and byte-identical after write_cells."""
    with zipfile.ZipFile(xlsm_with_extlst) as zf:
        original_bytes = zf.read("customXml/item1.xml")

    write_cells(xlsm_with_extlst, "Sheet1", [(1, "A", "MODIFIED")])

    with zipfile.ZipFile(xlsm_with_extlst) as zf:
        assert "customXml/item1.xml" in zf.namelist(), "customXml part must be copied"
        after_bytes = zf.read("customXml/item1.xml")

    assert after_bytes == original_bytes, "customXml/item1.xml must be byte-identical"


def test_round_trip_preserves_shared_strings_xml(minimal_xlsm: Path) -> None:
    """xl/sharedStrings.xml must be byte-identical after write_cells (not touched)."""
    with zipfile.ZipFile(minimal_xlsm) as zf:
        names = zf.namelist()
        if "xl/sharedStrings.xml" not in names:
            pytest.skip("fixture has no sharedStrings.xml")
        original = zf.read("xl/sharedStrings.xml")

    write_cells(minimal_xlsm, "Sheet1", [(1, "A", "ROUNDTRIP")])

    with zipfile.ZipFile(minimal_xlsm) as zf:
        after = zf.read("xl/sharedStrings.xml")

    assert after == original, "sharedStrings.xml must not be modified"


def test_round_trip_preserves_namespaces(xlsm_with_extlst: Path) -> None:
    """xmlns:x14, xmlns:mc, mc:Ignorable must survive after write_cells."""
    write_cells(xlsm_with_extlst, "Sheet1", [(1, "A", "NS_TEST")])

    with zipfile.ZipFile(xlsm_with_extlst) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        xml = zf.read(sheet_names[0])

    root = etree.fromstring(xml)
    ns_map = root.nsmap

    assert NS_X14 in ns_map.values(), f"xmlns:x14 must be declared; got {ns_map}"
    assert NS_MC in ns_map.values(), f"xmlns:mc must be declared; got {ns_map}"
    ignorable = root.get(f"{{{NS_MC}}}Ignorable")
    assert ignorable is not None, "mc:Ignorable attribute must survive"


# ---------------------------------------------------------------------------
# Step 4: sheet name resolution
# ---------------------------------------------------------------------------


def test_resolves_sheet_name_to_correct_xml(tmp_path: Path) -> None:
    """Japanese sheet name is resolved to the correct xl/worksheets/sheetN.xml."""
    xlsm = tmp_path / "jp_sheet.xlsm"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "工数集計"
    ws.cell(row=1, column=1, value="ORIGINAL")
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    with zipfile.ZipFile(xlsm, "w", compression=zipfile.ZIP_DEFLATED) as dst:
        with zipfile.ZipFile(io.BytesIO(buf.read())) as src:
            for info in src.infolist():
                dst.writestr(info, src.read(info.filename))

    write_cells(xlsm, "工数集計", [(1, "A", "日本語テスト")])

    wb2 = openpyxl.load_workbook(xlsm, data_only=True)
    assert wb2["工数集計"]["A1"].value == "日本語テスト"


# ---------------------------------------------------------------------------
# Step 5: atomic write — original intact on failure
# ---------------------------------------------------------------------------


def test_atomic_failure_keeps_original_intact(minimal_xlsm: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    """If os.replace raises during finalization, the original file is untouched."""
    original_bytes = minimal_xlsm.read_bytes()

    def _fail_replace(src: Any, dst: Any) -> None:
        raise OSError("simulated replace failure")

    monkeypatch.setattr("os.replace", _fail_replace)

    with pytest.raises(OSError, match="simulated replace failure"):
        write_cells(minimal_xlsm, "Sheet1", [(1, "A", "SHOULD_NOT_PERSIST")])

    # Original must be untouched
    assert minimal_xlsm.read_bytes() == original_bytes

    # Temp file must be cleaned up
    tmp = minimal_xlsm.with_suffix(".tmp.xlsm")
    assert not tmp.exists(), "tmp file must be cleaned up on failure"


# ---------------------------------------------------------------------------
# Step 6: type encoding
# ---------------------------------------------------------------------------


def test_write_cells_bool_encoding(minimal_xlsm: Path) -> None:
    write_cells(minimal_xlsm, "Sheet1", [(1, "A", True)])
    with zipfile.ZipFile(minimal_xlsm) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        xml = zf.read(sheet_names[0])
    root = etree.fromstring(xml)
    rows = root.findall(f".//{{{NS}}}row")
    row1 = next(r for r in rows if r.get("r") == "1")
    a1 = next(c for c in row1.findall(f"{{{NS}}}c") if c.get("r") == "A1")
    assert a1.get("t") == "b"
    assert a1.find(f"{{{NS}}}v").text == "1"


def test_write_cells_date_as_serial_number(minimal_xlsm: Path) -> None:
    """datetime.date is stored as Excel serial number (numeric cell, no t=)."""
    d = datetime.date(2026, 5, 14)
    write_cells(minimal_xlsm, "Sheet1", [(1, "A", d)])
    with zipfile.ZipFile(minimal_xlsm) as zf:
        sheet_names = [n for n in zf.namelist() if n.startswith("xl/worksheets/sheet")]
        xml = zf.read(sheet_names[0])
    root = etree.fromstring(xml)
    rows = root.findall(f".//{{{NS}}}row")
    row1 = next(r for r in rows if r.get("r") == "1")
    a1 = next(c for c in row1.findall(f"{{{NS}}}c") if c.get("r") == "A1")
    # Must be numeric (no t= attribute or t not 's'/'inlineStr')
    assert a1.get("t") not in ("s", "inlineStr", "b"), "date must be stored as number"
    v = a1.find(f"{{{NS}}}v")
    assert v is not None
    serial = int(v.text)
    assert serial > 0, "date serial must be positive"
