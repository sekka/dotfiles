"""Read rows from an Excel sheet for one-way pull."""
from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula


def _resolve_value(structural: Any, computed: Any) -> Any:
    """Return the appropriate cell value, preferring computed over formula strings."""
    if isinstance(structural, str) and structural.startswith("="):
        return computed
    if isinstance(structural, (ArrayFormula, DataTableFormula)):
        return computed
    return structural


def read_rows(
    workbook_path: Path,
    *,
    sheet: str,
    data_start_row: int,
    columns: list[str],
    id_column: str = "A",
) -> list[dict[str, Any]]:
    """Read consecutive rows starting at data_start_row, stopping at the first empty id cell.

    Returns a list of dicts keyed by column letter (e.g. {"A": "T-001", "B": "..."}).
    """
    wb_struct = openpyxl.load_workbook(workbook_path, data_only=False)
    ws_struct = wb_struct[sheet]
    wb_calc = openpyxl.load_workbook(workbook_path, data_only=True)
    ws_calc = wb_calc[sheet]
    id_idx = column_index_from_string(id_column)
    rows: list[dict[str, Any]] = []
    row = data_start_row
    while True:
        id_value = ws_struct.cell(row=row, column=id_idx).value
        if id_value is None or id_value == "":
            break
        row_data: dict[str, Any] = {}
        for col in columns:
            col_idx = column_index_from_string(col)
            structural = ws_struct.cell(row=row, column=col_idx).value
            computed = ws_calc.cell(row=row, column=col_idx).value
            row_data[col] = _resolve_value(structural, computed)
        rows.append(row_data)
        row += 1
    return rows
