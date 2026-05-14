"""Migration utilities for bootstrap id assignment."""
from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string

from lib import xlsm_writer as _xlsm_writer


def compute_id_assignments(
    existing_ids: list[str | None], prefix: str = "T-"
) -> list[tuple[int, str]]:
    """Return (index, new_id) pairs for entries whose id is empty.

    Parses existing non-empty ids to find the current maximum N, then assigns
    prefix+{N+1:03d}, prefix+{N+2:03d}, ... to empty slots in order.
    Non-matching ids (garbage, different prefix) are skipped when computing max.
    """
    max_n = 0
    pattern = re.compile(r"^" + re.escape(prefix) + r"(\d+)$")
    for v in existing_ids:
        if v is None:
            continue
        s = str(v).strip()
        m = pattern.match(s)
        if m:
            max_n = max(max_n, int(m.group(1)))

    assignments: list[tuple[int, str]] = []
    counter = max_n + 1
    for idx, v in enumerate(existing_ids):
        s = str(v).strip() if v is not None else ""
        if s == "":
            assignments.append((idx, f"{prefix}{counter:03d}"))
            counter += 1
    return assignments


def scan_rows(
    workbook_path: Path,
    *,
    sheet: str,
    data_start_row: int,
    id_column: str,
    all_columns: list[str],
) -> list[dict[str, Any]]:
    """Scan all rows starting at data_start_row until a fully-blank row.

    Returns a list of dicts (one per row) with keys from all_columns plus "row"
    (1-based row number). Rows where every column is None/empty are treated as
    the end sentinel and scanning stops.
    """
    keep_vba = workbook_path.suffix.lower() == ".xlsm"
    wb = openpyxl.load_workbook(workbook_path, keep_vba=keep_vba, data_only=True)
    ws = wb[sheet]

    col_indices = {col: column_index_from_string(col) for col in all_columns}
    id_idx = column_index_from_string(id_column)

    results: list[dict[str, Any]] = []
    row_num = data_start_row
    max_row = ws.max_row or data_start_row

    while row_num <= max_row:
        row_data: dict[str, Any] = {}
        for col, idx in col_indices.items():
            row_data[col] = ws.cell(row=row_num, column=idx).value
        # stop at fully-blank row (all tracked columns are None/"")
        if all(
            (v is None or str(v).strip() == "") for v in row_data.values()
        ):
            break
        row_data["row"] = row_num
        results.append(row_data)
        row_num += 1

    return results


def write_id_cells(
    workbook_path: Path,
    *,
    sheet: str,
    id_column: str,
    assignments: list[tuple[int, str]],
    row_map: list[int],
) -> None:
    """Write id values to the specified rows in the workbook.

    assignments: list of (index_in_row_map, new_id)
    row_map: list of 1-based row numbers corresponding to scanned rows
    """
    updates = [(row_map[idx], id_column, new_id) for idx, new_id in assignments]
    _xlsm_writer.write_cells(workbook_path, sheet, updates)
